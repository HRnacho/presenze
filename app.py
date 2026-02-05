from flask import Flask, render_template, request, jsonify, session, redirect, url_for
from datetime import datetime, timedelta
import json
import os
import calendar
from functools import wraps

app = Flask(__name__)
app.secret_key = 'direzionelavoro-presenze-2025-secret-key'

DATA_FILE = os.path.join('data', 'presenze.json')

USERS = {
    'gianluca': {'password': 'direzione2025', 'nome': 'Gianluca Bittoni'},
    'ignacio': {'password': 'direzione2025', 'nome': 'Ignacio Sorcaburu Ciglieri'},
    'simone': {'password': 'direzione2025', 'nome': 'Simone Mascellari'}
}

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def save_data(data):
    os.makedirs('data', exist_ok=True)
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('calendario'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if username in USERS and USERS[username]['password'] == password:
            session['username'] = username
            session['nome_completo'] = USERS[username]['nome']
            return redirect(url_for('calendario'))
        else:
            return render_template('login.html', error='Credenziali non valide')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/calendario')
@login_required
def calendario():
    now = datetime.now()
    year = int(request.args.get('year', now.year))
    month = int(request.args.get('month', now.month))
    data = load_data()
    utenti = [{'username': u, 'nome': i['nome']} for u, i in USERS.items()]
    return render_template('calendario.html', year=year, month=month, utenti=utenti,
                         current_user=session['username'], nome_completo=session['nome_completo'])

@app.route('/api/presenze/<year>/<month>')
@login_required
def get_presenze(year, month):
    data = load_data()
    key = f"{year}-{month.zfill(2)}"
    return jsonify(data.get(key, {}))

@app.route('/api/presenza', methods=['POST'])
@login_required
def salva_presenza():
    try:
        info = request.json
        data = load_data()
        date_obj = datetime.strptime(info['date'], '%Y-%m-%d')
        key = date_obj.strftime('%Y-%m')
        
        if key not in data:
            data[key] = {}
        if info['username'] not in data[key]:
            data[key][info['username']] = {}
        
        data[key][info['username']][info['date']] = {
            'tipo': info['tipo'],
            'ore_lavorate': float(info.get('ore_lavorate', 0)),
            'ore_assenza': float(info.get('ore_assenza', 0)),
            'note': info.get('note', ''),
            'updated_at': datetime.now().isoformat()
        }
        
        save_data(data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/api/elimina_presenza', methods=['POST'])
@login_required
def elimina_presenza():
    try:
        info = request.json
        data = load_data()
        date_obj = datetime.strptime(info['date'], '%Y-%m-%d')
        key = date_obj.strftime('%Y-%m')
        if key in data and info['username'] in data[key] and info['date'] in data[key][info['username']]:
            del data[key][info['username']][info['date']]
            save_data(data)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': 'Non trovata'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/esporta/<year>/<month>')
@login_required
def esporta_excel(year, month):
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment
    
    # Carica il template Excel
    template_path = os.path.join('templates', 'Foglio_presenze_UDINE_Dicembre_2025.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # STEP 1: Unmerge TUTTE le celle merged
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))
    
    # STEP 2: Aggiorna il mese in RIGA 1 (non tocchiamo, lasciamo PERIODO)
    # Non modifichiamo la riga 1
    
    month_names = ['gennaio', 'febbraio', 'marzo', 'aprile', 'maggio', 'giugno',
                   'luglio', 'agosto', 'settembre', 'ottobre', 'novembre', 'dicembre']
    
    # Carica i dati delle presenze
    data = load_data()
    key = f"{year}-{month.zfill(2)}"
    presenze_mese = data.get(key, {})
    
    num_days = calendar.monthrange(int(year), int(month))[1]
    
    # Mappa username -> righe nel template
    # Gianluca: Ord=6, Str=7, Ass=8, Giust=9
    # Ignacio: Ord=10, Str=11, Ass=12, Giust=13
    # Simone: Ord=14, Str=15, Ass=16, Giust=17
    user_rows = {
        'gianluca': {'ord': 6, 'str': 7, 'ass': 8, 'giust': 9},
        'ignacio': {'ord': 10, 'str': 11, 'ass': 12, 'giust': 13},
        'simone': {'ord': 14, 'str': 15, 'ass': 16, 'giust': 17}
    }
    
    # Festivi italiani 2025-2026
    holidays = {
        1: [1, 6],
        2: [],
        3: [],
        4: [20, 21],
        5: [1],
        6: [2],
        7: [],
        8: [15],
        9: [],
        10: [],
        11: [1],
        12: [8, 25, 26]
    }
    
    current_month_holidays = holidays.get(int(month), [])
    
    # STEP 3: Popola i dati per ogni dipendente
    # Colonna J = giorno 1, K = giorno 2, ... AN = giorno 31
    # J è la colonna numero 10
    
    for username, rows in user_rows.items():
        if username not in presenze_mese:
            continue
            
        presenze_user = presenze_mese[username]
        
        # Per ogni giorno del mese
        for day in range(1, num_days + 1):
            # Colonna J = 10, quindi giorno 1 = colonna 10, giorno 2 = colonna 11, etc.
            col_index = 9 + day  # J=10, quindi 9+1=10
            col_letter = openpyxl.utils.get_column_letter(col_index)
            
            date_str = f"{year}-{month.zfill(2)}-{str(day).zfill(2)}"
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            
            is_weekend = date_obj.weekday() in [5, 6]
            is_holiday = day in current_month_holidays
            
            # Weekend e festivi in rosso
            if is_weekend or is_holiday:
                red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                
                ws[f'{col_letter}{rows["ord"]}'].fill = red_fill
                ws[f'{col_letter}{rows["ord"]}'].value = None
                
                ws[f'{col_letter}{rows["str"]}'].fill = red_fill
                ws[f'{col_letter}{rows["str"]}'].value = None
                
                ws[f'{col_letter}{rows["ass"]}'].fill = red_fill
                ws[f'{col_letter}{rows["ass"]}'].value = None
                
                ws[f'{col_letter}{rows["giust"]}'].fill = red_fill
                ws[f'{col_letter}{rows["giust"]}'].value = None
                
                continue
            
            # Dati presenza
            if date_str in presenze_user:
                presenza = presenze_user[date_str]
                tipo = presenza.get('tipo', 'presenza')
                ore_lavorate = presenza.get('ore_lavorate', 0)
                ore_assenza = presenza.get('ore_assenza', 0)
                
                ore_ordinarie = max(0, min(8, ore_lavorate) - ore_assenza)
                ore_straordinari = max(0, ore_lavorate - 8) if ore_lavorate > 8 else 0
                
                # Ord
                cell_ord = ws[f'{col_letter}{rows["ord"]}']
                if ore_ordinarie > 0:
                    cell_ord.value = f"{ore_ordinarie:.2f}".replace('.', ',')
                    cell_ord.alignment = Alignment(horizontal='center')
                else:
                    cell_ord.value = None
                
                # Str
                cell_str = ws[f'{col_letter}{rows["str"]}']
                if ore_straordinari > 0:
                    cell_str.value = f"{ore_straordinari:.2f}".replace('.', ',')
                    cell_str.alignment = Alignment(horizontal='center')
                else:
                    cell_str.value = None
                
                # Ass
                cell_ass = ws[f'{col_letter}{rows["ass"]}']
                if ore_assenza > 0:
                    cell_ass.value = f"{ore_assenza:.2f}".replace('.', ',')
                    cell_ass.alignment = Alignment(horizontal='center')
                else:
                    cell_ass.value = None
                
                # Giust
                cell_giust = ws[f'{col_letter}{rows["giust"]}']
                if tipo != 'presenza' and ore_assenza > 0:
                    codice_map = {
                        'ferie': 'FERIE',
                        'rol': 'ROL',
                        'malattia': 'MALATTIA',
                        'permesso': 'PERMESSO'
                    }
                    cell_giust.value = codice_map.get(tipo, '')
                    cell_giust.alignment = Alignment(horizontal='center')
                else:
                    cell_giust.value = None
    
    # STEP 4: Ri-mergia le celle necessarie
    # Riga 1
    ws.merge_cells('A1:E1')  # PERIODO
    
    # Riga 3: intestazioni
    ws.merge_cells('A3:D3')   # Cognome
    ws.merge_cells('E3:H3')   # Nome
    ws.merge_cells('I3:AN3')  # PRESTAZIONI PER CIASCUNA GIORNATA
    
    # Cognome e Nome per ogni dipendente
    # Gianluca (righe 6-7)
    ws.merge_cells('A6:D7')   # Cognome
    ws.merge_cells('E6:H7')   # Nome
    
    # Ignacio (righe 10-11)
    ws.merge_cells('A10:D11') # Cognome
    ws.merge_cells('E10:H11') # Nome
    
    # Simone (righe 14-15)
    ws.merge_cells('A14:D15') # Cognome
    ws.merge_cells('E14:H15') # Nome
    
    # Salva
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Foglio_presenze_UDINE_{month_names[int(month)-1].title()}_{year}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    os.makedirs('data', exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5000)
